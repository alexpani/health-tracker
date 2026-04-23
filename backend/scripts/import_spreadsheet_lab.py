"""Importa lo storico analisi da uno spreadsheet xlsx nel dominio Lab Results.

Formato atteso (spec §9):
- Riga 1 = header date in colonne B, C, D, … (formato gg/mm/aa, gg/mm/aaaa,
  o qualsiasi cosa che openpyxl riconosca come datetime).
- Colonna A = nomi analiti (matching via `lab_analyte_aliases` case-insensitive).
- Riga opzionale "Note" (colonna A) → popola `lab_panels.notes` per ogni colonna.
- Celle: valore numerico (con virgola italiana) o testo qualitativo; se contiene
  un'unità inline (es. "3,02 pg/ml"), viene estratta in `unit_raw`.

Uso:
    python -m scripts.import_spreadsheet_lab --file storico.xlsx [--sheet Analisi] [--dry-run|--commit]

Default: --dry-run. Produce `import_report.tsv` accanto al file di input.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session as default_session_factory
from app.models.lab import LabPanel, LabResult


# ---------------------------------------------------------------------------
# Parsing cella
# ---------------------------------------------------------------------------

_NUMBER_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")
# es. "3,02 pg/ml", "0.78 ng/ml", "90" → ("90", None); "assente" → ("assente", None)
_VALUE_UNIT_RE = re.compile(
    r"^\s*(?P<value>-?\d+(?:[.,]\d+)?)\s*(?P<unit>[^\s].*)?$"
)


@dataclass
class CellParsed:
    value_numeric: Decimal | None = None
    value_text: str | None = None
    unit_raw: str | None = None


def parse_cell(raw: Any) -> CellParsed | None:
    """Ritorna None se la cella è vuota / non interpretabile."""
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        return CellParsed(value_numeric=Decimal(str(raw)))
    s = str(raw).strip()
    if not s:
        return None

    # Se è tutto numero (senza unità inline)
    if _NUMBER_RE.match(s):
        try:
            return CellParsed(value_numeric=Decimal(s.replace(",", ".")))
        except InvalidOperation:
            return CellParsed(value_text=s)

    # Prova "<numero> <unità>"
    m = _VALUE_UNIT_RE.match(s)
    if m:
        try:
            val = Decimal(m.group("value").replace(",", "."))
            unit = (m.group("unit") or "").strip() or None
            return CellParsed(value_numeric=val, unit_raw=unit)
        except InvalidOperation:
            pass

    # Fallback: testo qualitativo
    return CellParsed(value_text=s)


def parse_header_date(raw: Any) -> date | None:
    """Supporta Excel-native datetimes e stringhe gg/mm/aa, gg/mm/aaaa, ISO."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Report (dry-run)
# ---------------------------------------------------------------------------

@dataclass
class RowReport:
    row_idx: int
    raw_name: str
    analyte_id: int | None
    analyte_slug: str | None
    match_kind: str  # 'exact' | 'miss'
    columns_with_value: int


@dataclass
class PanelReport:
    col_letter: str
    test_date: date | None
    notes: str | None
    values_count: int
    unmapped_count: int


@dataclass
class ImportReport:
    panels: list[PanelReport] = field(default_factory=list)
    rows: list[RowReport] = field(default_factory=list)

    def unmapped_rows(self) -> list[RowReport]:
        return [r for r in self.rows if r.analyte_id is None and r.columns_with_value > 0]


def write_tsv_report(report: ImportReport, out_path: Path) -> None:
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["section", "key", "detail1", "detail2", "detail3"])
        for p in report.panels:
            w.writerow([
                "panel", p.col_letter,
                p.test_date.isoformat() if p.test_date else "(no-date)",
                f"values={p.values_count}",
                f"unmapped={p.unmapped_count}",
            ])
        for r in report.rows:
            w.writerow([
                "row", str(r.row_idx), r.raw_name,
                r.analyte_slug or f"<no-match>",
                f"values_across_panels={r.columns_with_value}",
            ])


# ---------------------------------------------------------------------------
# Core import
# ---------------------------------------------------------------------------

async def _lookup_analyte_id(db: AsyncSession, raw_name: str) -> tuple[int | None, str | None]:
    """Match esatto (case-insensitive) su `lab_analyte_aliases`.
    Ritorna `(analyte_id, slug)` oppure `(None, None)`."""
    row = (await db.execute(
        text(
            "SELECT al.analyte_id, a.slug FROM lab_analyte_aliases al "
            "JOIN lab_analytes a ON a.id = al.analyte_id "
            "WHERE LOWER(al.alias) = LOWER(:n) LIMIT 1"
        ),
        {"n": raw_name},
    )).first()
    if row is None:
        return None, None
    return row[0], row[1]


async def import_spreadsheet(
    xlsx_path: Path,
    sheet_name: str | None,
    commit: bool,
    session_factory: Any | None = None,
) -> ImportReport:
    """session_factory iniettabile nei test (default: app.database.async_session)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    if ws is None:
        raise RuntimeError(f"Sheet '{sheet_name}' non trovato")

    # Riga 1 = date (da colonna B in poi)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    panel_cols: list[tuple[int, str, date | None]] = []
    for cell in header_row[1:]:
        d = parse_header_date(cell.value)
        if d is None:
            # Salta colonne senza data valida (evita di creare panel spazzatura)
            continue
        panel_cols.append((cell.column, cell.column_letter, d))

    # Trova riga "Note" (prima colonna = "Note", case-insensitive)
    note_cells_by_col: dict[int, str | None] = {c: None for c, _, _ in panel_cols}
    data_rows: list[tuple[int, str, dict[int, Cell]]] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if not row:
            continue
        name_cell = row[0]
        raw_name = str(name_cell.value).strip() if name_cell.value is not None else ""
        if not raw_name:
            continue
        cells_by_col = {c.column: c for c in row[1:]}
        if raw_name.lower() in ("note", "notes"):
            for col_idx in note_cells_by_col:
                cell = cells_by_col.get(col_idx)
                note_cells_by_col[col_idx] = (
                    str(cell.value).strip() if cell is not None and cell.value is not None else None
                )
            continue
        data_rows.append((name_cell.row, raw_name, cells_by_col))

    report = ImportReport()
    factory = session_factory or default_session_factory

    async with factory() as db:
        # Pre-match tutti gli analiti (case-insensitive) in un colpo solo
        analyte_map: dict[str, tuple[int | None, str | None]] = {}
        for row_idx, raw_name, _ in data_rows:
            if raw_name not in analyte_map:
                analyte_map[raw_name] = await _lookup_analyte_id(db, raw_name)

        # Costruisci i panel report + conta valori per riga/colonna
        col_values_count: dict[int, int] = {c: 0 for c, _, _ in panel_cols}
        col_unmapped_count: dict[int, int] = {c: 0 for c, _, _ in panel_cols}
        row_values_count: dict[int, int] = {}

        parsed_rows: list[tuple[int, str, int | None, str | None, dict[int, CellParsed]]] = []
        for row_idx, raw_name, cells_by_col in data_rows:
            aid, slug = analyte_map[raw_name]
            parsed: dict[int, CellParsed] = {}
            for col_idx, col_letter, _ in panel_cols:
                cell = cells_by_col.get(col_idx)
                p = parse_cell(cell.value if cell is not None else None)
                if p is None:
                    continue
                parsed[col_idx] = p
                col_values_count[col_idx] += 1
                if aid is None:
                    col_unmapped_count[col_idx] += 1
            row_values_count[row_idx] = len(parsed)
            parsed_rows.append((row_idx, raw_name, aid, slug, parsed))
            report.rows.append(RowReport(
                row_idx=row_idx, raw_name=raw_name,
                analyte_id=aid, analyte_slug=slug,
                match_kind="exact" if aid is not None else "miss",
                columns_with_value=len(parsed),
            ))

        for col_idx, col_letter, d in panel_cols:
            report.panels.append(PanelReport(
                col_letter=col_letter, test_date=d,
                notes=note_cells_by_col.get(col_idx),
                values_count=col_values_count[col_idx],
                unmapped_count=col_unmapped_count[col_idx],
            ))

        if not commit:
            return report

        # --- Commit path ---
        # Tutto in una transazione: flush graduale ma commit unico finale.
        now = datetime.now(timezone.utc)
        for col_idx, col_letter, d in panel_cols:
            if col_values_count[col_idx] == 0:
                # Evita di creare panel vuoti
                continue
            panel = LabPanel(
                test_date=d,
                lab_name=None,
                specimen_types=["blood"],  # default MVP; refattorizzabile
                status="confirmed",
                confirmed_at=now,
                notes=note_cells_by_col.get(col_idx),
            )
            db.add(panel)
            await db.flush()

            for row_idx, raw_name, aid, _slug, parsed in parsed_rows:
                p = parsed.get(col_idx)
                if p is None:
                    continue
                db.add(LabResult(
                    panel_id=panel.id,
                    analyte_id=aid,
                    raw_name=raw_name,
                    value_numeric=p.value_numeric,
                    value_text=p.value_text,
                    unit_raw=p.unit_raw,
                    needs_review=(aid is None),
                    out_of_range=None,
                ))
        await db.commit()

    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Import storico lab da xlsx")
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--sheet", default=None)
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", default=True)
    mode.add_argument("--commit", action="store_true")
    args = parser.parse_args()

    if not args.file.exists():
        print(f"[errore] file non trovato: {args.file}", file=sys.stderr)
        return 2

    commit = bool(args.commit)
    print(f"[mode] {'COMMIT' if commit else 'DRY-RUN'}")
    print(f"[file] {args.file}  sheet={args.sheet or '(active)'}")

    report = asyncio.run(import_spreadsheet(args.file, args.sheet, commit))

    out_path = args.file.with_name("import_report.tsv")
    write_tsv_report(report, out_path)

    total_panels = len(report.panels)
    total_rows = len(report.rows)
    unmapped = len(report.unmapped_rows())
    total_values = sum(p.values_count for p in report.panels)

    print(f"[summary] panels={total_panels}  analiti righe={total_rows}  "
          f"valori celle={total_values}  analiti non mappati={unmapped}")
    if unmapped:
        print("\nAnaliti da rivedere (aggiungi alias o crea analita):")
        for r in report.unmapped_rows():
            print(f"  - riga {r.row_idx}: {r.raw_name!r}  "
                  f"({r.columns_with_value} valori)")
    print(f"\n[report] {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
